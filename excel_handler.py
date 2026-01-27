# -*- coding: utf-8 -*-
"""
Excel处理模块：实现Excel文件的创建和写入功能
"""

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from config import FieldMapping, Constants
import logging
import sys
import os


# 配置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class ExcelHandler:
    def __init__(self, template_file=None):
        """
        初始化Excel处理器
        :param template_file: Excel模板文件路径（可选）
        """
        self.template_file = template_file
        self.wb = None
        self.ws = None  # 主工作表
        
    def create_new_workbook(self):
        """
        创建新的Excel工作簿
        :return: 是否成功创建
        """
        try:
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.title = "财务数据分析"
            logger.info("成功创建新的Excel工作簿")
            return True
        except Exception as e:
            logger.error(f"创建Excel工作簿失败: {e}")
            return False
    
    def create_excel_template(self, template_file):
        """
        自动创建Excel模板文件
        :param template_file: 模板文件路径
        :return: 是否成功创建
        """
        try:
            # 创建新的Excel工作簿
            wb = Workbook()
            
            # 创建财务数据分析工作表
            ws_main = wb.active
            ws_main.title = "财务数据分析"
            
            # 设置列宽
            ws_main.column_dimensions['A'].width = 20
            ws_main.column_dimensions['B'].width = 15
            ws_main.column_dimensions['C'].width = 15
            ws_main.column_dimensions['D'].width = 15
            ws_main.column_dimensions['E'].width = 15
            ws_main.column_dimensions['F'].width = 15
            
            # 设置标题
            ws_main.merge_cells('A1:F1')
            ws_main['A1'] = "公司财务数据分析模板"
            ws_main['A1'].font = Font(size=16, bold=True)
            ws_main['A1'].alignment = Alignment(horizontal='center', vertical='center')
            
            # 设置基础信息区
            ws_main['A2'] = '当前股价'
            ws_main['B2'] = 0
            ws_main['B2'].number_format = '0.00'
            ws_main['C2'] = '元'
            
            ws_main['D2'] = '总市值'
            ws_main['E2'] = 0
            ws_main['E2'].number_format = '0.00'
            ws_main['F2'] = '亿元'
            
            ws_main['A3'] = '所属行业'
            ws_main['B3'] = ''
            
            ws_main['D3'] = '市盈率 PE'
            ws_main['E3'] = 0
            ws_main['E3'].number_format = '0.00'
            ws_main['F3'] = '倍'
            
            ws_main['A4'] = '市净率 PB'
            ws_main['B4'] = 0
            ws_main['B4'].number_format = '0.00'
            ws_main['C4'] = '倍'
            
            ws_main['D4'] = '5年复合增长率'
            ws_main['E4'] = 0
            ws_main['E4'].number_format = '0.00'
            ws_main['F4'] = '%'
            
            # 设置计算逻辑说明
            ws_main['A5'] = '净利润含金量计算逻辑: 经营活动产生的现金流量净额 / 归母净利润 * 100%'
            ws_main['A5'].font = Font(size=10, italic=True)
            ws_main['A5'].alignment = Alignment(horizontal='left', vertical='center')
            
            # 设置年份表头
            ws_main['A6'] = '指标'
            ws_main['B6'] = '2024'
            ws_main['C6'] = '2023'
            ws_main['D6'] = '2022'
            ws_main['E6'] = '2021'
            ws_main['F6'] = '2020'
            
            # 设置表头格式
            for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                cell = ws_main[f'{col}6']
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # 设置财务指标行
            indicator_labels = {
                7: '营业总收入（亿元）',
                8: '营业收入增长率（%）',
                9: '归母净利润（亿元）',
                10: '净利润增长率（%）',
                11: '净资产收益率 ROE（%）',
                12: '毛利率（%）',
                13: '净利率（%）',
                14: '净利润含金量（%）',
                15: '资产负债率（%）',
                16: '每股经营现金流（元）',
                17: '每股净资产（元）',
                18: '每股收益 EPS（元）'
            }
            
            for row, label in indicator_labels.items():
                ws_main[f'A{row}'] = label
                ws_main[f'A{row}'].font = Font(bold=True)
                ws_main[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
            
            # 创建情况表工作表
            ws_situation = wb.create_sheet(title="情况表")
            ws_situation['A1'] = '公司基本情况'
            ws_situation['A1'].font = Font(size=14, bold=True)
            ws_situation['A3'] = '主营业务'
            ws_situation['B3'] = ''
            ws_situation['A4'] = '行业地位'
            ws_situation['B4'] = ''
            
            # 保存模板文件
            wb.save(template_file)
            wb.close()
            logger.info(f"成功创建Excel模板文件: {template_file}")
            return True
        except Exception as e:
            logger.error(f"创建Excel模板文件失败: {e}")
            return False
    
    def open_template(self):
        """
        打开模板文件，如果模板文件不存在则自动创建
        :return: 是否成功打开
        """
        try:
            template_file = "2025年五年投资收益计算表-模版.xlsx"
            
            # 检查模板文件是否存在
            if not os.path.exists(template_file):
                logger.warning(f"模板文件不存在，正在自动创建: {template_file}")
                if not self.create_excel_template(template_file):
                    logger.error("创建模板文件失败，使用默认工作簿")
                    return self.create_new_workbook()
            
            # 打开模板文件
            self.wb = load_workbook(template_file)
            self.ws = self.wb.active
            
            # 检查是否存在情况表工作表
            if "情况表" in self.wb.sheetnames:
                self.ws_situation = self.wb["情况表"]
            else:
                self.ws_situation = self.wb.create_sheet(title="情况表")
            
            # 检查是否存在计算表工作表
            if "投资收益计算" in self.wb.sheetnames:
                self.ws_calculation = self.wb["投资收益计算"]
            else:
                self.ws_calculation = self.wb.create_sheet(title="投资收益计算")
            
            logger.info(f"成功打开Excel模板文件: {template_file}")
            return True
        except Exception as e:
            logger.error(f"打开模板文件失败: {e}")
            # 如果打开失败，创建新的工作簿
            return self.create_new_workbook()
    
    def setup_excel_structure(self, company_info, stock_code):
        """
        设置Excel表格结构，按照Row 1-20的核心指标要求
        :param company_info: 公司信息字典
        :param stock_code: 股票代码
        :return: 是否成功设置
        """
        try:
            # 设置列宽
            self.ws.column_dimensions['A'].width = 20
            self.ws.column_dimensions['B'].width = 15
            self.ws.column_dimensions['C'].width = 15
            self.ws.column_dimensions['D'].width = 15
            self.ws.column_dimensions['E'].width = 15
            self.ws.column_dimensions['F'].width = 15
            
            # Row 1: 标题（A1:F1合并）
            stock_name = company_info.get('stock_name', '')
            title = f"{stock_name} {stock_code} 财务数据分析"
            self.ws.merge_cells('A1:F1')
            self.ws['A1'] = title
            self.ws['A1'].font = Font(size=16, bold=True)
            self.ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            
            # Row 2: 当前股价、总市值
            self.ws['A2'] = '当前股价'
            self.ws['B2'] = company_info.get('current_price', 0)
            self.ws['B2'].number_format = '0.00'
            self.ws['C2'] = '元'
            
            self.ws['D2'] = '总市值'
            self.ws['E2'] = company_info.get('total_market_cap', 0)
            self.ws['E2'].number_format = '0.00'
            self.ws['F2'] = '亿元'
            
            # Row 3: 所属行业、市盈率
            self.ws['A3'] = '所属行业'
            self.ws['B3'] = company_info.get('industry', '')
            
            self.ws['D3'] = '市盈率 PE'
            self.ws['E3'] = company_info.get('pe_ratio', 0)
            self.ws['E3'].number_format = '0.00'
            self.ws['F3'] = '倍'
            
            # Row 4: 市净率、5年复合增长率
            self.ws['A4'] = '市净率 PB'
            self.ws['B4'] = company_info.get('pb_ratio', 0)
            self.ws['B4'].number_format = '0.00'
            self.ws['C4'] = '倍'
            
            self.ws['D4'] = '5年复合增长率'
            
            # 写入5年复合增长率值
            cagr_5y = company_info.get('cagr_5y', 0)
            if cagr_5y == 'N/A':
                self.ws['E4'] = 'N/A'
            else:
                # 转换为百分比形式（乘以100），因为F4已经有%单位
                self.ws['E4'] = cagr_5y * 100
            self.ws['E4'].number_format = '0.00'
            self.ws['F4'] = '%'
            
            # Row 5: 计算逻辑说明
            self.ws['A5'] = '净利润含金量计算逻辑: 经营活动产生的现金流量净额 / 归母净利润 * 100%'
            self.ws['A5'].font = Font(size=10, italic=True)
            self.ws['A5'].alignment = Alignment(horizontal='left', vertical='center')
            
            # Row 6: 年份表头
            self.ws['A6'] = '指标'
            self.ws['B6'] = '2024'
            self.ws['C6'] = '2023'
            self.ws['D6'] = '2022'
            self.ws['E6'] = '2021'
            self.ws['F6'] = '2020'
            
            # 设置表头格式
            for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                cell = self.ws[f'{col}6']
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # Row 7-20: 财务指标（带单位标注）
            indicator_labels = {
                7: '营业总收入（亿元）',
                8: '营业收入增长率（%）',
                9: '归母净利润（亿元）',
                10: '净利润增长率（%）',
                11: '净资产收益率 ROE（%）',
                12: '毛利率（%）',
                13: '净利率（%）',
                14: '净利润含金量（%）',
                15: '资产负债率（%）',
                16: '每股经营现金流（元）',
                17: '每股净资产（元）',
                18: '每股收益 EPS（元）'
            }
            
            for row, label in indicator_labels.items():
                self.ws[f'A{row}'] = label
                self.ws[f'A{row}'].font = Font(bold=True)
                self.ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
            
            logger.info("Excel表格结构设置完成")
            return True
            
        except Exception as e:
            logger.error(f"Excel表格结构设置失败: {e}")
            return False
    
    def update_header(self, company_info, stock_code):
        """
        更新基础信息区（A1 - F6）
        :param company_info: 公司信息字典
        :param stock_code: 股票代码
        :return: 是否成功更新
        """
        try:
            # A1:F1 合并单元格：股票名称 + 股票代码 + 2025年五年投资收益计算表
            stock_name = company_info.get('stock_name', '')
            title = f"{stock_name} {stock_code} 2025年五年投资收益计算表"
            try:
                self.ws_calculation.cell(row=1, column=1).value = title
            except Exception as e:
                logger.warning(f"无法写入标题（可能是合并单元格）: {e}")
            
            # B2: 当前股价（元）
            current_price = company_info.get('current_price', 0)
            try:
                self.ws_calculation.cell(row=2, column=2).value = current_price
                self.ws_calculation.cell(row=2, column=2).number_format = '0.00'
            except Exception as e:
                logger.warning(f"无法写入当前股价: {e}")
            
            # D2: 总市值（亿元）
            total_market_cap = company_info.get('total_market_cap', 0)
            try:
                self.ws_calculation.cell(row=2, column=4).value = total_market_cap
                self.ws_calculation.cell(row=2, column=4).number_format = '0.00'
            except Exception as e:
                logger.warning(f"无法写入总市值: {e}")
            
            # B3: 所属行业
            industry = company_info.get('industry', '')
            try:
                self.ws_calculation.cell(row=3, column=2).value = industry
            except Exception as e:
                logger.warning(f"无法写入所属行业: {e}")
            
            # D3: 静态市盈率 PE
            pe_ratio = company_info.get('pe_ratio', 0)
            try:
                self.ws_calculation.cell(row=3, column=4).value = pe_ratio
                self.ws_calculation.cell(row=3, column=4).number_format = '0.00'
            except Exception as e:
                logger.warning(f"无法写入市盈率: {e}")
            
            # B4: 市净率 PB
            pb_ratio = company_info.get('pb_ratio', 0)
            try:
                self.ws_calculation.cell(row=4, column=2).value = pb_ratio
                self.ws_calculation.cell(row=4, column=2).number_format = '0.00'
            except Exception as e:
                logger.warning(f"无法写入市净率: {e}")
            
            # D4: 5年复合增长率（预测值）- 从company_info读取
            cagr_5y = company_info.get('cagr_5y', 0)
            try:
                if cagr_5y == 'N/A':
                    self.ws_calculation.cell(row=4, column=4).value = 'N/A'
                else:
                    self.ws_calculation.cell(row=4, column=4).value = cagr_5y
                    self.ws_calculation.cell(row=4, column=4).number_format = '0.00%'
            except Exception as e:
                logger.warning(f"无法写入5年复合增长率: {e}")
            
            logger.info("基础信息区更新完成")
            return True
        except Exception as e:
            logger.error(f"基础信息区更新失败: {e}")
            return False
    
    def write_financial_indicators(self, financial_indicators):
        """
        写入财务指标数据
        :param financial_indicators: 财务指标数据
        :return: 是否成功写入
        """
        try:
            for indicator, year_mapping in FieldMapping.FINANCIAL_INDICATORS.items():
                for year, cell in year_mapping.items():
                    year_int = int(year)
                    if year_int in financial_indicators and indicator in financial_indicators[year_int]:
                        value = financial_indicators[year_int][indicator]
                        
                        # 异常处理：遇到"N/A"、None或空字符串时，跳过写入
                        if value is None or value == "" or value == "N/A" or (isinstance(value, str) and value.strip() == "N/A"):
                            logger.warning(f"跳过写入无效值: {indicator} {year} = {value}")
                            continue
                        
                        # 检查是否为错误值
                        if value in Constants.ERROR_VALUES.values():
                            self.ws_calculation[cell] = value
                        else:
                            # 根据指标类型设置数值格式
                            if indicator.endswith('_rate') or indicator == 'roe' or indicator == 'asset_liability_ratio':
                                # 百分比类型指标，value已经是百分比形式（如15.32），需要转换为小数（0.1532）
                                if isinstance(value, (int, float)):
                                    # 将百分比转换为小数形式
                                    decimal_value = value / 100
                                    self.ws_calculation[cell] = decimal_value
                                    # 设置单元格格式为百分比
                                    self.ws_calculation[cell].number_format = '0.00%'
                                else:
                                    # 非数值类型，直接写入
                                    self.ws_calculation[cell] = value
                            else:
                                # 数值类型指标
                                if isinstance(value, (int, float)):
                                    self.ws_calculation[cell] = value
                                    # 设置单元格格式为常规数字格式
                                    self.ws_calculation[cell].number_format = 'General'
                                else:
                                    # 非数值类型，直接写入
                                    self.ws_calculation[cell] = value
            
            logger.info("财务指标数据写入完成")
            return True
        except Exception as e:
            logger.error(f"财务指标数据写入失败: {e}")
            return False
    
    def write_company_info(self, company_info):
        """
        写入公司信息
        :param company_info: 公司信息
        :return: 是否成功写入
        """
        try:
            for field, cell in FieldMapping.COMPANY_INFO.items():
                if field in company_info:
                    try:
                        # 尝试直接写入
                        self.ws_calculation[cell] = company_info[field]
                    except Exception as cell_error:
                        # 如果写入失败，可能是合并单元格，跳过
                        logger.warning(f"单元格 {cell} 写入失败（可能是合并单元格），跳过写入 {field}: {cell_error}")
                        continue
            
            logger.info("公司信息写入完成")
            return True
        except Exception as e:
            logger.error(f"公司信息写入失败: {e}")
            return False
    
    def write_valuation_indicators(self, valuation_indicators):
        """
        写入估值指标
        :param valuation_indicators: 估值指标
        :return: 是否成功写入
        """
        try:
            for field, cell in FieldMapping.VALUATION_INDICATORS.items():
                if field in valuation_indicators:
                    self.ws_calculation[cell] = valuation_indicators[field]
            
            logger.info("估值指标写入完成")
            return True
        except Exception as e:
            logger.error(f"估值指标写入失败: {e}")
            return False
    
    def write_situation_data(self, situation_data):
        """
        写入情况表数据
        :param situation_data: 情况表数据
        :return: 是否成功写入
        """
        try:
            for field, cell in FieldMapping.SITUATION_SHEET.items():
                if field in situation_data:
                    self.ws_situation[cell] = situation_data[field]
            
            logger.info("情况表数据写入完成")
            return True
        except Exception as e:
            logger.error(f"情况表数据写入失败: {e}")
            return False
    
    def write_all_data(self, processed_data):
        """
        写入所有数据
        :param processed_data: 处理后的所有数据
        :return: 是否成功写入
        """
        try:
            # 写入财务指标
            if 'financial_indicators' in processed_data:
                self.write_financial_indicators(processed_data['financial_indicators'])
            
            # 写入公司信息
            if 'company_info' in processed_data:
                self.write_company_info(processed_data['company_info'])
            
            # 写入估值指标
            if 'valuation_indicators' in processed_data:
                self.write_valuation_indicators(processed_data['valuation_indicators'])
            
            # 写入情况表数据（使用默认值）
            situation_data = {
                'business_main_operations': "数据来源于东方财富网",
                'industry_position': "根据财务数据自动生成"
            }
            self.write_situation_data(situation_data)
            
            logger.info("所有数据写入完成")
            return True
        except Exception as e:
            logger.error(f"数据写入失败: {e}")
            return False
    
    def write_multicolumn_data(self, financial_data, stock_code):
        """
        使用多列并行填充方式写入财务数据
        :param financial_data: 按年份组织的财务数据字典
        :param stock_code: 股票代码
        :return: 是否成功写入
        """
        try:
            # 设置Excel表格结构（包含基础信息区）
            if 'company_info' in financial_data:
                self.setup_excel_structure(financial_data['company_info'], stock_code)
            
            # 写入财务指标区（Row 7-20）
            for indicator, config in FieldMapping.INDICATOR_MAP.items():
                row = config['row']
                
                # 遍历年份，填充到对应列
                for year, column in FieldMapping.YEAR_COLUMN_MAP.items():
                    if year in financial_data and indicator in financial_data[year]:
                        value = financial_data[year][indicator]
                        
                        try:
                            # 异常处理：遇到"N/A"、None或空字符串时，跳过写入
                            if value is None or value == "" or value == "N/A" or (isinstance(value, str) and value.strip() == "N/A"):
                                logger.warning(f"跳过写入无效值: {indicator} {year} = {value}")
                                continue
                            
                            # 使用精准定位写入数据
                            cell = self.ws.cell(row=row, column=column)
                            
                            # 写入值
                            cell.value = value
                            
                            # 设置单元格格式
                            if config.get('display_format') == 'percentage':
                                cell.number_format = '0.00%'
                            else:
                                cell.number_format = '0.00'
                            
                            # 设置对齐方式
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                            
                            # 检查净利润含金量，低于80%标记为红色
                            if indicator == 'net_profit_cash_ratio' and 'warning_threshold' in config:
                                if isinstance(value, (int, float)) and value < config['warning_threshold']:
                                    cell.font = Font(color='FF0000', bold=True)
                                    cell.fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
                            
                        except Exception as cell_error:
                            logger.warning(f"写入单元格 ({row}, {column}) 失败: {cell_error}")
                            continue
            
            logger.info("多列数据写入完成")
            return True
            
        except Exception as e:
            logger.error(f"多列数据写入失败: {e}")
            return False
    
    def save_file(self, output_file):
        """
        保存Excel文件
        :param output_file: 输出文件路径
        :return: 是否成功保存
        """
        try:
            # 如果文件已存在，尝试删除
            import os
            if os.path.exists(output_file):
                try:
                    os.remove(output_file)
                    logger.info(f"已删除旧文件: {output_file}")
                except Exception as e:
                    logger.warning(f"无法删除旧文件: {e}")
            
            # 保存文件
            self.wb.save(output_file)
            logger.info(f"成功保存Excel文件: {output_file}")
            return True
        except PermissionError as e:
            logger.error(f"保存Excel文件失败，权限被拒绝: {e}")
            logger.error(f"请确保文件 {output_file} 未被其他程序打开")
            return False
        except Exception as e:
            logger.error(f"保存Excel文件失败: {e}")
            return False
    
    def save_to_memory(self):
        """
        将Excel文件保存到内存流
        :return: BytesIO对象
        """
        try:
            import io
            output = io.BytesIO()
            self.wb.save(output)
            output.seek(0)
            logger.info("成功将Excel文件保存到内存流")
            return output
        except Exception as e:
            logger.error(f"保存Excel到内存流失败: {e}")
            return None
    
    def close(self):
        """
        关闭Excel文件
        """
        try:
            if self.wb:
                self.wb.close()
                logger.info("Excel文件已关闭")
        except Exception as e:
            logger.error(f"关闭Excel文件失败: {e}")

if __name__ == "__main__":
    # 测试Excel处理功能
    from config import Constants
    
    # 创建测试数据
    test_data = {
        'financial_indicators': {
            2020: {
                'roe': 31.41,
                'net_profit_growth_rate': 13.33,
                'revenue_growth_rate': 10.29,
                'net_profit': 466.97,
                'non_recurring_profit': 457.64,
                'operating_cash_flow': 451.18,
                'dividend_rate': 1.82,
                'asset_liability_ratio': 19.53
            },
            2021: {
                'roe': 29.89,
                'net_profit_growth_rate': 12.03,
                'revenue_growth_rate': 11.71,
                'net_profit': 523.99,
                'non_recurring_profit': 513.16,
                'operating_cash_flow': 621.10,
                'dividend_rate': 1.91,
                'asset_liability_ratio': 19.08
            }
        },
        'company_info': {
            'stock_code': 'SH600519',
            'stock_name': '贵州茅台',
            'industry': '酿酒行业',
            'total_market_cap': 21000.0
        },
        'valuation_indicators': {
            'pe': 25.67,
            'pb': 10.23,
            'eps': 50.12
        }
    }
    
    # 测试Excel写入
    handler = ExcelHandler("2025年五年投资收益计算表-模版.xlsx")
    if handler.open_template():
        handler.write_all_data(test_data)
        handler.save_file("测试输出.xlsx")
        handler.close()
        print("Excel写入测试完成")
